#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim: ai ts=4 sts=4 et sw=4 nu

import os
import re
import sys
import json
import pathlib
import subprocess
from contextlib import contextmanager

import gdown
from slugify import slugify
from openpyxl import load_workbook
from zimscraperlib.video.presets import VideoWebmHigh
from zimscraperlib.video.encoding import reencode


DATA_DIR = pathlib.Path("data").resolve()
FILENAMES_MAP_PATH = pathlib.Path("filenames.map")
FILENAMES_MAP = {}


@contextmanager
def cwd(path):
    oldpwd = os.getcwd()
    os.chdir(path)
    try:
        yield
    finally:
        os.chdir(oldpwd)


def update_map(gid, fname):
    FILENAMES_MAP[gid] = fname
    with open(FILENAMES_MAP_PATH, "w") as fh:
        json.dump(FILENAMES_MAP, fh, indent=4)


def create_collection_for(ws, overwrite=False):
    name = slugify(ws.title)
    print(f"{ws.title} -> {name}")

    data_dir = DATA_DIR / name
    data_dir.mkdir(exist_ok=True)

    items = []

    for c_title, c_desc, c_auth, c_url in ws.iter_rows(min_row=2, max_col=4):
        if c_title.value is None:
            break

        print(c_title.value)

        # retrieve google drive ID from field
        # in-doc url fmt: https://docs.google.com/uc?id=[FILE_ID]&export=download
        # actual url fmt: https://drive.google.com/file/d/<id>/view?usp=sharing
        gid = re.search(r"([a-zA-Z0-9\_\-]+)/view", c_url.value).groups()[-1]

        # retrieve expected filename from map
        fname = FILENAMES_MAP.get(gid)
        fpath = data_dir / fname if fname else None
        # download if we dont already have it
        if not fname or (fpath and not fpath.exists()) or overwrite:
            fpath = pathlib.Path(
                gdown.download(
                    f"https://drive.google.com/uc?id={gid}", output=f"{data_dir}/"
                )
            )
            orig_fpath = data_dir / fpath.relative_to(data_dir).name
            fpath = orig_fpath.with_suffix(".webm")
            # reencode video to webm
            reencode(
                orig_fpath,
                fpath,
                VideoWebmHigh().to_ffmpeg_args(),
                delete_src=True,
            )

            update_map(gid, fpath.relative_to(data_dir).name)
        else:
            print("  Skipping download")

        # update collection
        items.append(
            {
                "title": c_title.value,
                "description": c_desc.value,
                "authors": c_auth.value,
                "files": [str(fpath.name)],
            }
        )

    # dump collection to file
    with open(data_dir / "collection.json", "w") as fh:
        json.dump(items, fh, indent=4)

    # create ZIP
    zip_path = data_dir.parent / f"{name}.zip"
    if not zip_path.exists() or overwrite:
        zip_path.unlink(missing_ok=True)
        with cwd(data_dir):
            args = [
                "zip",
                "-r",
                "-0",
                "-T",
                f"{zip_path.resolve()}",
                "collection.json",
            ] + [f.name for f in data_dir.iterdir()]
            subprocess.run(args)


def convert(path):
    global FILENAMES_MAP
    print(f"Working off {path}")

    # create data-dir if it doesn't exist
    DATA_DIR.mkdir(exist_ok=True)

    # load filenames map if it exists
    if FILENAMES_MAP_PATH.exists():
        with open(FILENAMES_MAP_PATH, "r") as fh:
            FILENAMES_MAP = json.load(fh)

    wb = load_workbook(filename=str(path))
    for name in wb.sheetnames:
        create_collection_for(wb[name])


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} xlsx-file")
        raise SystemExit(1)
    convert(pathlib.Path(sys.argv[1]).expanduser().resolve())
